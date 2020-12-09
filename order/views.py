from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response
from django.http.response import JsonResponse
from rest_framework.parsers import JSONParser 
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework import status
from datetime import datetime
from django.utils import timezone
from master_data.models import Customer,Station
from model_DTO import responseDTO
from model_DTO.base_DTO import base_DTO
from order.models import File,Order
from master_data.models import Part,Package,RouterMaster
import openpyxl 
import csv
import json
from django.views.decorators.csrf import csrf_exempt
from model_DTO.validateError import validateError,validateErrorList
from model_DTO.responseDTO import responseDTO
import os 
from .serializers import FileSerializer,File_Serializer_DTO,File_list_Serializer_DTO,validateErrorSerializer,validateErrorSerializerList,OrderSerializer,Order_list_Serializer_DTO,Order_transaction_Serializer,Order_transaction_Serializer_DTO,Order_transaction_list_Serializer_DTO
from rest_framework.permissions import IsAuthenticated
from annoying.functions import get_object_or_None
from django.db.models.expressions import RawSQL
from django.db.models import Q
import uuid
from rest_framework.authentication import TokenAuthentication
from rest_framework.decorators import permission_classes
from rest_framework.decorators import authentication_classes
import json
from .model.order_history import order_history
from .model.order_transaction import Order_transaction
from decimal import Decimal
from functools import reduce
import operator, random
from django.db.models.functions import Lower
import math
from openpyxl.utils import get_column_letter
from app.helper.config.ConfigMessage import ConfigMessage
from app.helper.order_helper.OrderUploadHelper import OrderUploadHelper
from app.helper.workbook_helper.WorkbookHelper import WorkbookHelper
from app.helper.order_helper.OrderValidatedHelper import OrderValidatedHelper
from app.serializersMapping.SerializerMapping import SerializerMapping
from app.helper.order_helper.OrderManageHelper import OrderManageHelper
from app.helper.CSV_file_management.CSVFileManagement import CSVFileManagement
from app.helper.order_helper.OrderComfirmHelper import OrderComfirmHelper
from app.helper.file_management.FileManagement import FileManagement
from app.services.order_service.OrderService import OrderService
from app.helper.order_helper.OrderMissMatchHelper import OrderMissMatchHelper
from app.helper.order_helper.OrderUploadLogHelper import OrderUploadLogHelper
from app.helper.order_helper.OrderTransactionHelper import OrderTransactionHelper
from app.helper.config.ConfigPart import ConfigPart
from django.conf import settings


configMessage = ConfigMessage()
configPart = ConfigPart()
serializerMapping = SerializerMapping()

class FileUploadView(APIView):
    
    parser_class = (FileUploadParser,)
    today_date_str = datetime.utcnow().strftime("%y%m%d")
    letter_list = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
    validate_error_list = []
    order_csv_list = []
    order_csv_list_database = []
    order_list = []

    permission_classes=[IsAuthenticated]


    def post(self, request, *args, **kwargs):

        try:
           
            customer_code = request.POST.get("customer_code", "")
            project_code = request.POST.get("project_code", "")

            uploaded_part = FileManagement.validate_folder(configPart.configs.get("UPLOAD_ORDER_PART").data)
            uploaded_part = "media/" + uploaded_part + "/"
            self.order_csv_list_database = []
            self.order_csv_list = []

            file_serializer = FileSerializer(data=request.data)
            
            if file_serializer.is_valid():

                file_obj = file_serializer.save(
                        file_name = request.FILES['file'].name ,
                        file_size = request.FILES['file'].size,
                        updated_by = request.user.username,
                    )
                sheet_obj = WorkbookHelper.read_workbook(file_serializer.data['file'][1:])

                row_max_int = sheet_obj.max_row
                column_max_int = sheet_obj.max_column

                return_list = OrderUploadHelper.data_mapping_from_sheet(sheet_obj)
    
                orderValidatedHelper = OrderValidatedHelper(return_list[0],return_list[3],return_list[4])
                error_list = orderValidatedHelper.get_error_list()
            

                if len(error_list) > 0 :
                    
                    serializer = serializerMapping.mapping_serializer_list(validateErrorSerializerList,None,"Error",configMessage.configs.get("UPLOAD_FILE_MASSAGE_ERROR").data,None,None,sorted(error_list, key=lambda error: (error.row,error.column) ) )
                    File.objects.filter(file_no=file_obj.file_no).delete()
                    return Response(serializer.data, status=status.HTTP_200_OK)

                else :

                    orderManageHelper = OrderManageHelper(file_obj.file_no,return_list[0],return_list[5])
                    orderManageHelper.order_management()
                    name_csv_str = file_obj.file_no + "DatabaseCSV"
                    CSV_file_management_obj = CSVFileManagement(
                        name_csv_str,
                        uploaded_part,
                        '',";")
                    CSV_file_management_obj.covert_to_CSV_data_list(orderManageHelper.get_order_csv_list_database())
                    return_name_CSV_str = CSV_file_management_obj.genearete_CSV_file()

                    name_csv_str = file_obj.file_no
                    CSV_file_management_obj = CSVFileManagement(
                        name_csv_str,
                        uploaded_part,
                        '',',')
                    CSV_file_management_obj.covert_to_header(["Item No","Order ID","Part Number","Part Description","Supplier Name","Plant","Order Amount","Date"])
                    order_csv_list =  orderManageHelper.get_order_csv_list()
                    CSV_file_management_obj.covert_to_CSV_data_list(orderManageHelper.get_order_csv_list())
                    return_name_CSV_str = CSV_file_management_obj.genearete_CSV_file()

                    File.objects.filter(file_no=file_obj.file_no).update(order_count=len(order_csv_list))

                    serializer = serializerMapping.mapping_serializer_list(validateErrorSerializerList,None,"success",configMessage.configs.get("UPLOAD_FILE_MASSAGE_SUCCESSFUL").data,None,None,None )
                    return Response(serializer.data, status=status.HTTP_200_OK)
               
        except Exception as e:

            serializer = serializerMapping.mapping_serializer_list(validateErrorSerializerList,None,"Error",e,None,None,[] )

            return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_files(request):

    try: 

        if request.method == 'GET':
      
            file_list = File.objects.filter(updated_by= request.user.username, status = 1)

            if len(file_list) > 0 :

                csv_name = FileManagement.find_file(configPart.configs.get("UPLOAD_ORDER_PART").data,file_list[0].file_no+".csv")
                serializer = serializerMapping.mapping_serializer_list(File_list_Serializer_DTO,file_list,"success",None,csv_name,None,None )

            else :

                serializer = serializerMapping.mapping_serializer_list(File_list_Serializer_DTO,[],"success",None,None,None,None )

            return Response(serializer.data,  status=status.HTTP_200_OK)

    except Exception as e:
        
        serializer = serializerMapping.mapping_serializer_list(FileSerializer,None,"Error",e,None,None,None )

        return Response(serializer.data, status=status.HTTP_200_OK)
        

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def confirm(request):

    if request.method == 'GET':

        try: 
            
            orderComfirmHelper = OrderComfirmHelper(request.user.username)
            order_list = orderComfirmHelper.order_comfirm_manage()

            serializer = serializerMapping.mapping_serializer_list(Order_list_Serializer_DTO,order_list,"success", configMessage.configs.get("UPLOAD_ORDER_MASSAGE_SUCCESSFUL").data,None,None,None )        
      
            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:

            print(e)

            serializer = serializerMapping.mapping_serializer_list(FileSerializer,None,"Error",e,None,None,None )

            return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def not_confirm(request):

    if request.method == 'POST':

        try:

            file_deleted_obj = File.objects.filter(updated_by= request.user.username, status = 1)[0]
            file_no_str = file_deleted_obj.file_no
            file_file_str = file_deleted_obj.file

            order_deleted_obj = Order.objects.filter(file_no=file_no_str)

            file_deleted_obj.delete()
            order_deleted_obj.delete()
            
            serializer = serializerMapping.mapping_serializer_obj(File_Serializer_DTO,file_deleted_obj,"success", configMessage.configs.get("CANCEN_CONFIRM_ORDER_MASSAGE_SUCCESSFUL").data,None,None,None )        

            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:

            print(e)
            serializer = serializerMapping.mapping_serializer_obj(File_Serializer_DTO,None,"Error",e,None,None,None )
       
            return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def search_miss_match(request):

    if request.method == 'POST':

        try:
            
            customer_code = request.data['customer_selected']
            project_code = request.data['project_selected']
            supplier_code = request.data['supplier_selected']
            plant_code = request.data['plant_selected']
            start_date = request.data['start_date_selected']
            end_date = request.data['end_date_selected']

            CSV_part_str = FileManagement.validate_folder(configPart.configs.get("MISS_MATCH_ORDER_PART").data)
            CSV_part_generete_str = 'media/' + CSV_part_str + "/"

            orderService = OrderService()
            order_list =  orderService.search_miss_match(customer_code,project_code,supplier_code,plant_code,start_date,end_date)

            order_serializer_list = OrderMissMatchHelper.covert_data_list_to_serializer_list(order_list)

            name_csv_str = "OrderMissMatchCSV_" +datetime.now().strftime("%Y%m%d_%H%M%S")
            CSV_file_management_obj = CSVFileManagement(name_csv_str,CSV_part_generete_str,'',',')
            CSV_file_management_obj.covert_to_header([
                "File ID",
                "Order ID",
                "Supplier",
                "Plant",
                "Part No",
                "Due Date",
                "Order Qty",
                "Pakage No",
                "Pakage Qty",
                "Route&Qty",
                "Uploaded By",
                "Uploaded Date"])
            order_CSV_list = OrderMissMatchHelper.covert_data_list_to_CSV_list(order_list)
            CSV_file_management_obj.covert_to_CSV_data_list(order_CSV_list)
            return_name_CSV_str = CSV_file_management_obj.genearete_CSV_file()
           
            order_serializer = OrderSerializer(order_serializer_list, many=True)

            serializer = serializerMapping.mapping_serializer_list(
                Order_list_Serializer_DTO,
                order_serializer.data,
                "success", 
                "",
                CSV_part_str +"/"+ name_csv_str + ".csv",
                None,
                None )


            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:

            serializer = serializerMapping.mapping_serializer_list(Order_list_Serializer_DTO,None,"Error",e,None,None,None )

            return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def match_order(request):

    if request.method == 'POST':

        try :

            orderMissMatchHelper = OrderMissMatchHelper(request.user.username)
            orderMissMatchHelper.miss_match_management()

            serializer = serializerMapping.mapping_serializer_list(
                Order_list_Serializer_DTO,
                None,
                "success", 
                configMessage.configs.get("MATCH_ORDER_MASSAGE_SUCCESSFUL").data,
                None,
                None,
                None )

            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:

            serializer = serializerMapping.mapping_serializer_list(Order_list_Serializer_DTO,None,"Error",e,None,None,None )

            return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def search_pending_order(request):

    if request.method == 'POST':

        try:
            
            customer_code = request.data['customer_selected']
            project_code = request.data['project_selected']
            supplier_code = request.data['supplier_selected']
            plant_code = request.data['plant_selected']
            start_date = request.data['start_date_selected']
            end_date = request.data['end_date_selected']

            CSV_part_str = FileManagement.validate_folder(configPart.configs.get("PENDING_ORDER_PART").data)
            CSV_part_generete_str = 'media/' + CSV_part_str + "/"

            orderService = OrderService()
            order_list =  orderService.search_pending_order(
                customer_code,
                project_code,
                supplier_code,
                plant_code,
                start_date,
                end_date)

            order_serializer_list = OrderMissMatchHelper.covert_data_list_to_serializer_list(order_list)


            name_csv_str = "OrderPendingCSV_" +datetime.now().strftime("%Y%m%d_%H%M%S")
            CSV_file_management_obj = CSVFileManagement(name_csv_str,CSV_part_generete_str,'',',')
            CSV_file_management_obj.covert_to_header([
                "File ID",
                "Order ID",
                "Supplier",
                "Plant",
                "Part No",
                "Due Date",
                "Order Qty",
                "Pakage No",
                "Pakage Qty",
                "Route&Qty",
                "Uploaded By",
                "Uploaded Date"])
            order_CSV_list = OrderMissMatchHelper.covert_data_list_to_CSV_list(order_list)
            CSV_file_management_obj.covert_to_CSV_data_list(order_CSV_list)
            return_name_CSV_str = CSV_file_management_obj.genearete_CSV_file()
           
            order_serializer = OrderSerializer(order_serializer_list, many=True)

            serializer = serializerMapping.mapping_serializer_list(
                Order_list_Serializer_DTO,
                order_serializer.data,
                "success", 
                "",
                CSV_part_str + "/" + name_csv_str + ".csv",
                None,
                None )


            return Response(serializer.data, status=status.HTTP_200_OK)
          
        except Exception as e:

            serializer = serializerMapping.mapping_serializer_list(Order_list_Serializer_DTO,None,"Error",e,None,None,None )

            return Response(serializer.data, status=status.HTTP_200_OK)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def search_upload_order_log_file(request):

    if request.method == 'POST':

        try:
            

            customer_code = request.data['customer_selected']
            project_code = request.data['project_selected']
            start_date = request.data['start_date_selected']
            end_date = request.data['end_date_selected']

            CSV_part_str = FileManagement.validate_folder(configPart.configs.get("UPLOAD_LOG_ORDER_PART").data)
            CSV_part_generete_str = 'media/' + CSV_part_str + "/"

            orderService = OrderService()
            file_list =  orderService.search_upload_order_log_file(
                customer_code,
                project_code,
                start_date,
                end_date)

            serializer_list = OrderUploadLogHelper.covert_data_list_to_serializer_list(file_list)

            name_csv_str = "UloadOrderLogFileCSV_" +datetime.now().strftime("%Y%m%d_%H%M%S")
            CSV_file_management_obj = CSVFileManagement(name_csv_str,CSV_part_generete_str,'',',')
            CSV_file_management_obj.covert_to_header([
                "Customer Code",
                "Project",
                "File No",
                "Order Count",
                "Status",
                "Upload By",
                "Upload Date"])
            file_CSV_list = OrderUploadLogHelper.covert_data_list_to_CSV_list(file_list)
            CSV_file_management_obj.covert_to_CSV_data_list(file_CSV_list)
            return_name_CSV_str = CSV_file_management_obj.genearete_CSV_file()

            file_serializer = FileSerializer(serializer_list, many=True)
            serializer = serializerMapping.mapping_serializer_list(
                File_list_Serializer_DTO,
                serializer_list,
                "success", 
                "",
                CSV_part_str + "/" + name_csv_str + ".csv",
                None,
                None )


            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:

            serializer = serializerMapping.mapping_serializer_list(File_list_Serializer_DTO,None,"Error",e,None,None,None )

            return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET', 'POST', 'DELETE'])
def file_list(request):

    if request.method == 'GET':

        try : 

            file_list = File.objects.filter(status=2)
            file_serializer = FileSerializer(file_list, many=True)

            base_DTO_obj =  base_DTO()
            base_DTO_obj.serviceStatus = "success"
            base_DTO_obj.massage = "get file is success"
            base_DTO_obj.data_list = file_serializer.data

            file_list_Serializer_DTO_obj = File_list_Serializer_DTO(base_DTO_obj)


            return JsonResponse(file_list_Serializer_DTO_obj.data, safe=False)

        except Exception as e:

            base_DTO_obj =  base_DTO()
            base_DTO_obj.serviceStatus = "Error"
            base_DTO_obj.massage = e
            base_DTO_obj.data = None

            file_Serializer_DTO_reponse = File_Serializer_DTO(base_DTO_obj)

            return JsonResponse(file_Serializer_DTO_reponse.data, safe=False)
    
@api_view(['GET', 'POST', 'DELETE'])
def order_list(request):

    if request.method == 'GET':

        try : 

            order_list = Order.objects.filter(is_part_completed=True,is_route_completed=True)
            order_serializer = OrderSerializer(order_list, many=True)

            base_DTO_obj =  base_DTO()
            base_DTO_obj.serviceStatus = "success"
            base_DTO_obj.massage = "get order is success"
            base_DTO_obj.data_list = order_serializer.data

            order_list_Serializer_DTO_obj = Order_list_Serializer_DTO(base_DTO_obj)


            return JsonResponse(order_list_Serializer_DTO_obj.data, safe=False)

        except Exception as e:

            base_DTO_obj =  base_DTO()
            base_DTO_obj.serviceStatus = "Error"
            base_DTO_obj.massage = e
            base_DTO_obj.data = None

            order_Serializer_DTO_reponse = Order_Serializer_DTO(base_DTO_obj)

            return JsonResponse(order_Serializer_DTO_reponse.data, safe=False)

            if project_serializer_obj.is_valid():

                project_serializer_obj.save()

                base_DTO_obj =  base_DTO()
                base_DTO_obj.serviceStatus = "success"
                base_DTO_obj.massage = "project is saved"
                base_DTO_obj.data_list = Project.objects.all()

                project_list_serializer_DTO_reponse = Project_list_Serializer_DTO(base_DTO_obj)

                return JsonResponse(project_list_serializer_DTO_reponse.data, safe=False) 
            
            base_DTO_obj =  base_DTO()
            base_DTO_obj.serviceStatus = "Error"
            base_DTO_obj.massage = project_serializer_obj.errors
            base_DTO_obj.data = None

            project_Serializer_DTO_reponse = Project_Serializer_DTO(base_DTO_obj)

            return JsonResponse(project_Serializer_DTO_reponse.data,safe=False)

        except Exception as e:

            base_DTO_obj =  base_DTO()
            base_DTO_obj.serviceStatus = "Error"
            base_DTO_obj.massage = "someting is error"
            base_DTO_obj.data = None

            project_Serializer_DTO_reponse = Project_Serializer_DTO(base_DTO_obj)

            return JsonResponse(project_Serializer_DTO_reponse.data, safe=False)

@api_view(['POST'])
def search_order_transaction(request):

    if request.method == 'POST':

        try:

            customer_code = request.data['customer_selected']
            project_code = request.data['project_selected']
            start_date = request.data['start_date_selected']
            end_date = request.data['end_date_selected']
            file_no = request.data['file_selected']
            order_no = request.data['order_selected']
            supplier_code = request.data['supplier_selected']
            plant_code = request.data['plant_selected']

            CSV_part_str = FileManagement.validate_folder(configPart.configs.get("TRANSACTION_ORDER_PART").data)
            CSV_part_generete_str = 'media/' + CSV_part_str + "/"

            orderService = OrderService()
            order_list =  orderService.search_order_transaction(
                customer_code,
                project_code,
                file_no,
                order_no,
                supplier_code,
                plant_code,
                start_date,
                end_date)
            
            orderTransactionHelper = OrderTransactionHelper()
            order_transaction_list = orderTransactionHelper.transaction_management(order_list)
            
            name_csv_str = "OrderTransactionCSV_" +datetime.now().strftime("%Y%m%d_%H%M%S")
            CSV_file_management_obj = CSVFileManagement(name_csv_str,CSV_part_generete_str,'',',')
            CSV_file_management_obj.covert_to_header([
                "Action",
                "File ID",
                "Order ID",
                "Supplier",
                "Plant",
                "Part No",
                "Part Name",
                "Due Date",
                "Order Qty",
                "Package No",
                "Package Qty",
                "Route&Trip",
                "Uploaded By",
                "Uploaded Date"])
            # file_CSV_list = OrderUploadLogHelper.covert_data_list_to_CSV_list(file_list)
            CSV_file_management_obj.covert_to_CSV_data_list(orderTransactionHelper.order_CSV_list)
            return_name_CSV_str = CSV_file_management_obj.genearete_CSV_file()

            order_transaction_Serializer = Order_transaction_Serializer(order_transaction_list, many=True)
            serializer = serializerMapping.mapping_serializer_list(
                Order_transaction_list_Serializer_DTO,
                order_transaction_list,
                "success", 
                "",
                CSV_part_str + "/" +name_csv_str+".csv",
                None,
                None )

            return Response(serializer.data, status=status.HTTP_200_OK)

         
   

        except Exception as e:

            print(e)
            serializer = serializerMapping.mapping_serializer_list(File_Serializer_DTO,None,"Error",e,None,None,None )

            return Response(serializer.data, status=status.HTTP_200_OK)







        