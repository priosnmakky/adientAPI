from master_data.models import Station
from model_DTO.validateError import validateError,validateErrorList
from app.helper.config.ConfigMessage import ConfigMessage
from openpyxl.utils import get_column_letter
# from order.models import File
# from datetime import datetime\
configMessage = ConfigMessage()

class OrderValidatedHelper:

    supplier_workbook_list = []
    plant_workbook_list = [] 
    order_list = []
    error_list = []

    def __init__(self,order_list,supplier_workbook_list,plant_workbook_list):

        self.order_list = order_list
        self.supplier_workbook_list = supplier_workbook_list
        self.plant_workbook_list = plant_workbook_list
     
    def get_supplier_in_db(self) :

        supplier_db_list = set(Station.objects.filter(station_code__iregex=r'(' + '|'.join(self.supplier_workbook_list) + ')',station_type__iexact="SUPPLIER",is_active=True).values_list("station_code",flat=True))
        supplier_upper_db_list = set([s.upper() for s in supplier_db_list])

        return supplier_upper_db_list
    
    def get_supplier_in_workbook(self) :

        supplier_workbook_set = set([s.upper() for s in self.supplier_workbook_list])

        return supplier_workbook_set
    
    def get_supplier_codet_in_workbook(self) : 
        
        return set(self.get_supplier_in_workbook() - self.get_supplier_in_db())
    
    def get_plant_in_db(self) : 
        
        plant_db_list = Station.objects.filter(station_code__iregex=r'(' + '|'.join(self.plant_workbook_list) + ')',station_type__iexact="PLANT",is_active=True).values_list("station_code",flat=True)
        plant_upper_db_set = set([p.upper() for p in plant_db_list])

        return plant_upper_db_set
    
    def get_plant_in_workbook(self) :

        plant_workbook_set = set([p.upper() for p in self.plant_workbook_list])

        return plant_workbook_set
    
    def get_plant_codet_in_workbook(self) : 
        
        return set(self.get_plant_in_workbook() - self.get_plant_in_db())

    def validate_item_list(self) :
        
        item_list = [ x for x in self.order_list if x[2] == 1 and (not x[0].isnumeric() or x[0] == "") ]

        for item_error in item_list:

            validate_error_obj = validateError() 

            if item_error[0] is None or item_error[0] == "" :

                validate_error_obj.error = configMessage.configs.get("UPLOAD_ORDER_ITEM_REQUIRED").data
                validate_error_obj.row = item_error[1]
                validate_error_obj.column = get_column_letter(item_error[2])
                            
            elif not item_error[0].isnumeric() :

                validate_error_obj.error = configMessage.configs.get("UPLOAD_ORDER_ITEM_INTEGER").data
                validate_error_obj.row = item_error[1]
                validate_error_obj.column = get_column_letter(item_error[2])
            
            self.error_list.append(validate_error_obj)

        
        return self.error_list
    

    def validate_partNumber_list(self) :
        
        partNumber_list = [ x for x in self.order_list if x[2] == 2 and x[0] == "" ]

        for partNumber_error in partNumber_list:

            validate_error_obj = validateError() 

            if partNumber_error[0] is None or partNumber_error[0] == "" :

                validate_error_obj.error = configMessage.configs.get("UPLOAD_ORDER_PARTNUMBER_REQUIRED").data
                validate_error_obj.row = partNumber_error[1]
                validate_error_obj.column = get_column_letter(partNumber_error[2])
            
            self.error_list.append(validate_error_obj)

        
        return self.error_list
    
    def validate_partdescription_list(self) :
        
        partDescription_list = [ x for x in self.order_list if x[2] == 3 and x[0] == "" ]

        for partDescription_error in partDescription_list:

            validate_error_obj = validateError() 

            if partDescription_error[0] is None or partDescription_error[0] == "" :

                validate_error_obj.error = configMessage.configs.get("UPLOAD_ORDER_PARTDESCRIPTION_REQUIRED").data
                validate_error_obj.row = partDescription_error[1]
                validate_error_obj.column = get_column_letter(partDescription_error[2])
            
            self.error_list.append(validate_error_obj)

        
        return self.error_list
    

    def validate_supplier_list(self) :
        
        supplier_list = [ x for x in self.order_list if x[0].upper() in list(self.get_supplier_codet_in_workbook()) and x[2] == 4 ]

        for supplier_error in supplier_list:

            validate_error_obj = validateError() 

            station_len_int = len(Station.objects.filter(station_code__iexact= str(supplier_error[0]) ,station_type__iexact="PLANT",is_active=True))
           
            if supplier_error[0] is None or supplier_error[0] == "" :

                validate_error_obj.error = configMessage.configs.get("UPLOAD_ORDER_SUPPLIERCODE_REQUIRED").data
                validate_error_obj.row = supplier_error[1]
                validate_error_obj.column = get_column_letter(supplier_error[2])

            elif station_len_int == 0 :
    
                validate_error_obj.error = configMessage.configs.get("UPLOAD_ORDER_SUPPLIERCODE_DATABASE").data
                validate_error_obj.row = supplier_error[1]
                validate_error_obj.column = get_column_letter(supplier_error[2])

            self.error_list.append(validate_error_obj)
        
        return self.error_list

    
    def validate_plant_list(self) :
        
        plant_list = [ x for x in self.order_list if x[0].upper()  in list(self.get_plant_codet_in_workbook()) and x[2] == 5 ]
        # print(list(self.get_plant_codet_in_workbook()))
        for plant_error in plant_list:

            validate_error_obj = validateError() 
        
            station_len_int = len(Station.objects.filter(station_code__iexact= str(plant_error[0]),station_type__iexact="PLANT",is_active=True))
                            
            if plant_error[0] is None or plant_error[0] == "" :

                validate_error_obj.error = configMessage.configs.get("UPLOAD_ORDER_PLANTCODE_REQUIRED").data
                validate_error_obj.row = plant_error[1]
                validate_error_obj.column = get_column_letter(plant_error[2])
                            
            elif station_len_int == 0 :

                validate_error_obj.error = configMessage.configs.get("UPLOAD_ORDER_PLANTCODE_DATABASE").data
                validate_error_obj.row = plant_error[1]
                validate_error_obj.column = get_column_letter(plant_error[2]) 

            self.error_list.append(validate_error_obj)
        
        return self.error_list
    

    def validate_orderQry_list(self) :

        orderQry_list = [ x for x in self.order_list if x[2] > 5 and not x[0].isnumeric() ]

        for orderQry_error in orderQry_list:

            validate_error_obj = validateError() 

            if orderQry_error[0] is None or orderQry_error[0] == "" :

                validate_error_obj.error = configMessage.configs.get("UPLOAD_ORDER_QUANTITY_REQUIRED").data
                validate_error_obj.row = orderQry_error[1]
                validate_error_obj.column = get_column_letter(orderQry_error[2])

            elif not orderQry_error[0].isnumeric() :

                validate_error_obj.error = configMessage.configs.get("UPLOAD_ORDER_QUANTITY_INTEGER").data
                validate_error_obj.row = orderQry_error[1]
                validate_error_obj.column = get_column_letter(orderQry_error[2])
               
            self.error_list.append(validate_error_obj)
        
        return self.error_list



    def get_error_list(self) :
            
        self.error_list = []
        self.validate_item_list()
        self.validate_partNumber_list()
        self.validate_partdescription_list()
        self.validate_supplier_list()
        self.validate_plant_list()
        self.validate_orderQry_list()
        
        return self.error_list
    
    
    



    

    # @staticmethod
    # def generate_file_no(customer_id) :

    #     today_date_str = datetime.utcnow().strftime("%y%m%d")
    #     customer_code_str = Customer.objects.get(customer_code = customer_id ).customer_code
    #     customer_code_2dg_first_start_str = customer_code_str[0:2]
    #     file_count = File.objects.filter(file_no__contains= today_date_str ).count()
    #     print(file_count)
    #     file_no = customer_code_2dg_first_start_str + today_date_str + "{0:0=2d}".format(file_count + 1)

    #     return file_no
    
  
    
        