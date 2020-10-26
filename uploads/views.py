from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response
from django.http.response import JsonResponse
from rest_framework.parsers import JSONParser 
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework import status
from datetime import datetime
from django.utils import timezone
from customers.models import Customer
from model_DTO import responseDTO
from uploads.models import File
from order.models import Order
import openpyxl 
import csv
import json
from django.views.decorators.csrf import csrf_exempt
from model_DTO.validateError import validateError,validateErrorList
from model_DTO.responseDTO import responseDTO
import os 


from .serializers import FileSerializer,validateErrorSerializer,validateErrorSerializerList
from rest_framework.permissions import IsAuthenticated


class FileUploadView(APIView):
    parser_class = (FileUploadParser,)
    today_date_str = datetime.now().strftime("%y%m%d")
    letter_list = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]

    permission_classes=[IsAuthenticated]

    def generate_file_no(self,customer_id):


        customer_code_str = Customer.objects.get(pk=int(customer_id)).name
        customer_code_2dg_first_start_str = customer_code_str[0:2]
        file_count = File.objects.filter(file_no__contains= self.today_date_str ).count()
        file_count = file_count + 1
        if file_count < 10:
            file_count = "0" + str(file_count) 

        file_no = customer_code_2dg_first_start_str + self.today_date_str + str(file_count)

        return file_no

    def validated_cell(self,sheet_obj):

        row_max_int = sheet_obj.max_row
        column_max_int = sheet_obj.max_column

        validate_error_list = []

        print(row_max_int,column_max_int);

        for row_int in range(3, row_max_int + 1, 1):

            for column_int in range(1, column_max_int + 1, 1):
                
                order_cell_str = sheet_obj.cell(row=row_int, column=column_int).value

                if order_cell_str is not None :
                    if column_int == 1 and isinstance(order_cell_str, int) == False : 

                        validate_error_obj = validateError()                         
                        validate_error_obj.error = "Item no must be only number"
                        validate_error_obj.row = row_int
                        validate_error_obj.column = column_int                  
                        
                        validate_error_list.append(validate_error_obj)

                    if column_int == 4 and len(order_cell_str) > 6 :

                        validate_error_obj = validateError()                         
                        validate_error_obj.error = "Supplier Name must be over 6 letter"
                        validate_error_obj.row = row_int
                        validate_error_obj.column = column_int                             
                        
                        validate_error_list.append(validate_error_obj)
                
                    if column_int == 5 and len(order_cell_str) > 10 :
                        
                        validate_error_obj = validateError()                         
                        validate_error_obj.error = "Plant must be over 6 letter"
                        validate_error_obj.row = row_int
                        validate_error_obj.column = column_int                       
                        
                        validate_error_list.append(validate_error_obj)
                    
                    if column_int > 5 and isinstance(order_cell_str, int) == False :
                        
                        validate_error_obj = validateError()                         
                        validate_error_obj.error = "Item no must be only number"
                        validate_error_obj.row = row_int
                        validate_error_obj.column = column_int
                        print(row_int,column_int);
                        validate_error_list.append(validate_error_obj)

        return validate_error_list

    def get_order_csv_list(self,sheet_obj):
        row_max_int = sheet_obj.max_row
        column_max_int = sheet_obj.max_column

        validate_error_list = []

        order_csv_list = []
        order_csv_list.append(["Item No","Order Id","Part Description","Supplier Name","Plant","Order Amount","Date"])

        order_obj = Order.objects.filter(order_no__contains= self.today_date_str ).order_by("created_date")
        order_no_last_str = ""
        if len(order_obj) > 0 :
            order_no_last_str = order_obj[-1].order_no

        else :
            order_no_last_str = '0000000000'



        for row_int in range(3, row_max_int + 1, 1):
                
            order_row_list = []

            for column_int in range(6, column_max_int + 1, 1):

                order_cell_str = sheet_obj.cell(row=row_int, column=column_int).value
                if order_cell_str is not None and int(order_cell_str) > 0 :
                    order_no_str = self.generate_order_no(order_no_last_str)
                    order_row_list = [
                        sheet_obj.cell(row=row_int, column=1).value,
                        order_no_str,
                        sheet_obj.cell(row=row_int, column=3).value,
                        sheet_obj.cell(row=row_int, column=4).value,
                        sheet_obj.cell(row=row_int, column=5).value,
                        order_cell_str,
                        sheet_obj.cell(row=2, column=column_int).value

                    ]

                    order_no_last_str = order_no_str

                    order_csv_list.append(order_row_list)


        return order_csv_list

                    
    def generate_order_no(self,order_no_last_str):

        no_strt = self.today_date_str

        if order_no_last_str[0:6] == self.today_date_str :
            # print(order_no_last_str[7:])
            order_number_int = int(order_no_last_str[7:])

            if order_number_int == 999:
                # print("order_number_int")
                letter_index_int = self.letter_list.index(order_no_last_str[6]) 
                no_strt = no_strt + self.letter_list[letter_index_int + 1] + "001"

            else :
                order_number_int = order_number_int + 1
                # print(order_number_int)
                order_number_str = ""

                if order_number_int < 10:
                    order_number_str = "00" + str(order_number_int) 
                elif order_number_int < 100:
                    order_number_str = "0" + str(order_number_int)

                else :
                    order_number_str = str(order_number_int)
                    

                no_strt = no_strt +order_no_last_str[6]+ order_number_str
                # print(no_strt)

        else :

            no_strt = no_strt + "A001"


        return no_strt

                




        # today_date_str = self.today_date_str
        # order_no = today_date_str
        # order_obj_list = Order.objects.filter(order_no__contains= today_date_str ).order_by("created_date")
        # order_obj_count_int = order_obj_list.count()

        # if order_obj_count_int > 0 :
        #     last_order_no_obj = order_obj_list[order_obj_count_int-1].order_no
        #     letter_str = last_order_no_obj[6]
        #     letter_index_int = 0
        #     for i,letter in enumerate(letter_list):
        #         if letter == letter_str :
        #             index_int = i
            
        #     order_no_letter_count_int = order_obj_list.filter(order_no__contains= letter_list[letter_index_int]).count()

        #     if order_no_letter_count_int == 999 :

        #         order_no = order_no + letter_list[letter + 1] + "0001"

        #     else:

        #         order_no_sum_int = str(order_no_letter_count_int + 1)
        #         order_no_sum_str = ""
        #         if order_no_sum_int < 10:
        #             order_no_sum_str = "000" + str(file_count) 
        #         if order_no_sum_int < 100:
        #             order_no_sum_str = "00" + str(file_count)


        #         order_no = order_no + letter_str + order_no_sum_str

        # else:
        #     order_no = order_no + "A0001"

        # return order_no

    def post(self, request, *args, **kwargs):

        file_serializer = FileSerializer(data=request.data)
        
        if file_serializer.is_valid():

            file_no = self.generate_file_no(request.POST.get("customer_id", ""))

            file_serializer.save(file_no = file_no , 
                file_name = request.FILES['file'].name ,
                file_size = request.FILES['file'].size,
                customer_id = request.POST.get("customer_id", ""),
                project_id = request.POST.get("project_id", ""),
                status = False,
                created_by = request.user.username,
                created_date = datetime.now(tz=timezone.utc).strptime(datetime.now().strftime("%d-%m-%Y %H:%M:%S"),"%d-%m-%Y %H:%M:%S")
                )
            
            
            workbook_obj = openpyxl.load_workbook(file_serializer.data['file'][1:])
            sheet_obj = workbook_obj.active
            # print(len(self.validated_cell(file_serializer.data['file'])))
            validated_cell_list = self.validated_cell(sheet_obj)
            validated_cell_count_int = len(validated_cell_list)

            if validated_cell_count_int > 0:

                validateErrorListObj =  validateErrorList()
                validateErrorListObj.status = "error"
                validateErrorListObj.validateErrorList = validated_cell_list
                serializer = validateErrorSerializerList(validateErrorListObj)
                # # serializer. = "error"
                # serializer.validateErrorSerializerList = validateErrorSerializer(validated_cell_list, many=True)


                return Response(serializer.data, status=status.HTTP_201_CREATED)

            else:

                order_csv_list = self.get_order_csv_list(sheet_obj)

                with open("media/" + file_no + '.csv', 'w', newline='') as file:
                    writer = csv.writer(file)
                    writer.writerows(order_csv_list)

                File.objects.filter(pk=file_serializer.data['id']).update(order_count = len(order_csv_list))

                file_list = File.objects.filter(created_by= request.user.username, status = False)

                validateErrorListObj =  validateErrorList()
                validateErrorListObj.status = "success"
                validateErrorListObj.fileList = file_list

                serializer = validateErrorSerializerList(validateErrorListObj)

                return Response(serializer.data, status=status.HTTP_201_CREATED)
   

            # load file .xlsx for call data in file
            # workbook_obj = openpyxl.load_workbook(file_serializer.data['file'][1:])
            # sheet_obj = workbook_obj.active
            # row_max_int = sheet_obj.max_row
            # column_max_int = sheet_obj.max_column
            
            # order_csv_list = []
            # order_csv_list.append(["Item No","Part Number","Part Description","Supplier Name","Plant","Order Amount","Date"])

            # for row_int in range(3, row_max_int, 1):
                
            #     order_row_list = []

            #     for column_int in range(6, column_max_int, 1):

            #         order_cell_str = sheet_obj.cell(row=row_int, column=column_int).value
            #         if order_cell_str is not None and int(order_cell_str) > 0 :
            #             order_row_list = [
            #                 sheet_obj.cell(row=row_int, column=1).value,
            #                 sheet_obj.cell(row=row_int, column=2).value,
            #                 sheet_obj.cell(row=row_int, column=3).value,
            #                 sheet_obj.cell(row=row_int, column=4).value,
            #                 sheet_obj.cell(row=row_int, column=5).value,
            #                 order_cell_str,
            #                 sheet_obj.cell(row=2, column=column_int).value

            #                             ]
            #             order_csv_list.append(order_row_list)

                    
            # file_name_str = str(file_serializer.data['file']).split('/')[2].split(".")[0]

            # with open("media/" + file_name_str + '.csv', 'w', newline='') as file:
            #     writer = csv.writer(file)
            #     writer.writerows(order_csv_list)

            # File.objects.filter(pk=file_serializer.data['id']).update(order_count = len(order_csv_list))

            # responseDTO_ojb = responseDTO('success','upload file is successful',file_serializer.data)
            # responseDTO_ojb.status = 'success'
            # responseDTO_ojb.massage = 'upload file is successful'
            # responseDTO_ojb.data = file_serializer.data

#             serializer = CommentSerializer(comment)
# serializer.data

            # validateError.error = "asdasdasd"
            # serializer = validateErrorSerializer(validateError)

            # return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)



    




    
    # def generate_order_no(today_date_str):

    #     order_no = today_date_str;
    #     order_obj_list = Order.objects.filter(order_no__contains= today_date_str ).order_by(created_date)
    #     order_obj_count_int = order_obj_list.count()

    #     if order_obj_count_int > 0 :
    #         last_order_no_obj = order_obj_list[order_obj_count_int-1].order_no
    #         letter_str = last_order_no_obj[6]
    #         letter_index_int = 0
    #         for i,letter in enumerate(letter_list):
    #             if letter == letter_str :
    #                 index_int = i
            
    #         order_no_letter_count_int = order_obj_list.filter(order_no__contains= letter_list[letter_index_int]).count()

    #         if order_no_letter_count_int == 999 :

    #             order_no = order_no + letter_list[letter + 1] + "0001"

    #         else:

    #             order_no_sum_int = str(order_no_letter_count_int + 1)
    #             order_no_sum_str = ""
    #             if order_no_sum_int < 10:
    #                 order_no_sum_str = "000" + str(file_count) 
    #             if order_no_sum_int < 100:
    #                 order_no_sum_str = "00" + str(file_count)


    #             order_no = order_no + letter_str + order_no_sum_str

    #     else:
    #         order_no = order_no + "A0001"

    #     return order_no





        # order_obj_list = Order.objects.filter(order_no__contains= today_date_str ).order_by(created_date)
        #                             order_obj_count_int = order_obj_list.count()
        #                             if order_obj_count_int > 999:
        #                                 order_letter_str = Order.objects[order_obj_count_int-1].order_no[6]
        #                                 letter_idex_int = get_index_letter_method(order_letter_str)
        #                                 if letter_idex_int == 26:

        # for i,letter in enumerate(letter_list):

        #     if letter == last_letter

        #         return i

@api_view(['GET'])

def get(request):
    if request.method == 'GET':
        # print(request.user.username)
        file_list = File.objects.filter(created_by= request.user.username, status = False)

        file_serializer = FileSerializer(file_list, many=True)

        return JsonResponse(file_serializer.data, safe=False)


@api_view(['GET'])

def confirm(request):

    if request.method == 'GET':
        print(request.user.username)

        # File.objects.filter(created_by= request.user.username, status = False).update(status = True)
        file_list = File.objects.filter(created_by= request.user.username, status = False)

        print(file_list)
        with open("media/" + file_list[0].file_no + '.csv', newline='') as csvfile:
            spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
            cell_count_int = 0
            for row in spamreader:
                if cell_count_int > 0 :
                   print(row)
                cell_count_int = cell_count_int + 1

        file_serializer = FileSerializer(file_list, many=True)

        return JsonResponse(file_serializer.data, safe=False)



@api_view(['POST'])
def not_confirm(request):

    

        print(request.user)

        File.objects.filter(created_by= request.user.username, status = False).delete()
        file_list = File.objects.filter(created_by= request.user.username, status = False).values()

        file_serializer = FileSerializer(file_list, many=True)

        return JsonResponse(file_serializer.data, safe=False)

 
    # elif request.method == 'POST':
    #     order_data = JSONParser().parse(request)
    #     order_serializer = OrderSerializer(data=order_data)
    #     if order_serializer.is_valid():
    #         order_serializer.save()
    #         return JsonResponse(order_serializer.data, status=status.HTTP_201_CREATED) 
    #     return JsonResponse(order_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
